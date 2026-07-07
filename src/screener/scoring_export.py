import os
import sqlite3
import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

from src.screener.engine import load_screener_dataframe
from src.screener.presets import (
    quality_compounder,
    value_pick,
    growth_accelerator,
    dividend_champion,
    debt_free_blue_chip,
    turnaround_watch,
)


DB_PATH = "database/stock_analysis.db"
OUTPUT_PATH = "output/screener_output.xlsx"


# ============================================================
# P10 / P90 WINSORISED SCALING
# ============================================================

def winsor_scale(series, higher_is_better=True):
    """
    Cap values at P10 and P90 and scale to 0-100.
    """

    numeric = pd.to_numeric(
        series,
        errors="coerce"
    )

    valid = numeric.dropna()

    if valid.empty:
        return pd.Series(
            np.nan,
            index=series.index
        )

    p10 = valid.quantile(0.10)
    p90 = valid.quantile(0.90)

    if p90 == p10:
        result = pd.Series(
            50.0,
            index=series.index
        )

        result[numeric.isna()] = np.nan

        return result

    capped = numeric.clip(
        lower=p10,
        upper=p90
    )

    scaled = (
        (capped - p10)
        /
        (p90 - p10)
        * 100
    )

    if not higher_is_better:
        scaled = 100 - scaled

    return scaled


# ============================================================
# LOAD ADDITIONAL HISTORICAL DATA
# ============================================================

def load_historical_data():
    conn = sqlite3.connect(DB_PATH)

    try:
        pnl = pd.read_sql_query(
            """
            SELECT
                company_id,
                year,
                sales,
                net_profit,
                operating_profit,
                other_income,
                interest
            FROM profitandloss
            """,
            conn
        )

        bs = pd.read_sql_query(
            """
            SELECT
                company_id,
                year,
                equity_capital,
                reserves,
                borrowings
            FROM balancesheet
            """,
            conn
        )

        cf = pd.read_sql_query(
            """
            SELECT *
            FROM cashflow
            """,
            conn
        )

    finally:
        conn.close()

    return pnl, bs, cf


# ============================================================
# YEAR NORMALISATION
# ============================================================

def normalize_year(df):
    df = df.copy()

    df["year_num"] = pd.to_numeric(
        df["year"]
        .astype(str)
        .str.extract(
            r"((?:19|20)\d{2})"
        )[0],
        errors="coerce"
    )

    return df.dropna(
        subset=["year_num"]
    )


# ============================================================
# BUILD EXTRA SCORE METRICS
# ============================================================

def build_extra_metrics(base_df):
    pnl, bs, cf = load_historical_data()

    pnl = normalize_year(pnl)
    bs = normalize_year(bs)
    cf = normalize_year(cf)

    pnl = pnl.drop_duplicates(
        ["company_id", "year_num"],
        keep="last"
    )

    bs = bs.drop_duplicates(
        ["company_id", "year_num"],
        keep="last"
    )

    cf = cf.drop_duplicates(
        ["company_id", "year_num"],
        keep="last"
    )

    # --------------------------------------------------------
    # DETECT CASH FLOW COLUMN NAMES
    # --------------------------------------------------------

    cfo_candidates = [
        "operating_activity",
        "cash_from_operating_activity",
        "cash_from_operations"
    ]

    cfi_candidates = [
        "investing_activity",
        "cash_from_investing_activity"
    ]

    cfo_col = next(
        (
            col for col in cfo_candidates
            if col in cf.columns
        ),
        None
    )

    cfi_col = next(
        (
            col for col in cfi_candidates
            if col in cf.columns
        ),
        None
    )

    if cfo_col is None:
        raise KeyError(
            "Could not find CFO column in cashflow table."
        )

    if cfi_col is None:
        raise KeyError(
            "Could not find investing activity column."
        )

    # --------------------------------------------------------
    # CALCULATE FCF HISTORY
    # --------------------------------------------------------

    cf["fcf"] = (
        cf[cfo_col].fillna(0)
        +
        cf[cfi_col].fillna(0)
    )

    # --------------------------------------------------------
    # LATEST ROCE
    # --------------------------------------------------------

    roce = pnl.merge(
        bs,
        on=["company_id", "year_num"],
        how="inner",
        suffixes=("_pnl", "_bs")
    )

    roce["ebit"] = (
        roce["operating_profit"].fillna(0)
        +
        roce["other_income"].fillna(0)
        -
        roce["interest"].fillna(0)
    )

    roce["capital_employed"] = (
        roce["equity_capital"].fillna(0)
        +
        roce["reserves"].fillna(0)
        +
        roce["borrowings"].fillna(0)
    )

    roce["computed_roce_pct"] = np.where(
        roce["capital_employed"] > 0,
        roce["ebit"]
        /
        roce["capital_employed"]
        * 100,
        np.nan
    )

    latest_roce = (
        roce
        .sort_values(
            ["company_id", "year_num"]
        )
        .groupby("company_id")
        .tail(1)
        [
            [
                "company_id",
                "computed_roce_pct"
            ]
        ]
    )

    # --------------------------------------------------------
    # CFO / PAT 5-YEAR AVERAGE
    # --------------------------------------------------------

    cfo_pat_data = cf[
        [
            "company_id",
            "year_num",
            cfo_col
        ]
    ].merge(
        pnl[
            [
                "company_id",
                "year_num",
                "net_profit"
            ]
        ],
        on=["company_id", "year_num"],
        how="inner"
    )

    cfo_pat_records = []

    for company_id, group in cfo_pat_data.groupby(
        "company_id"
    ):
        group = group.sort_values(
            "year_num"
        ).tail(5)

        valid = group[
            group["net_profit"] != 0
        ].copy()

        if valid.empty:
            ratio = np.nan
        else:
            valid["ratio"] = (
                valid[cfo_col]
                /
                valid["net_profit"]
            )

            ratio = valid["ratio"].mean()

        cfo_pat_records.append(
            {
                "company_id": company_id,
                "cfo_pat_ratio_5yr": ratio
            }
        )

    cfo_pat_df = pd.DataFrame(
        cfo_pat_records
    )

    # --------------------------------------------------------
    # FCF CAGR 5-YEAR
    # --------------------------------------------------------

    fcf_records = []

    for company_id, group in cf.groupby(
        "company_id"
    ):
        group = group.sort_values(
            "year_num"
        )

        if len(group) < 6:
            value = np.nan
        else:
            latest = group.iloc[-1]
            target_year = int(
                latest["year_num"] - 5
            )

            old = group[
                group["year_num"]
                == target_year
            ]

            if old.empty:
                value = np.nan
            else:
                start = old.iloc[-1]["fcf"]
                end = latest["fcf"]

                if start > 0 and end > 0:
                    value = (
                        (
                            (end / start)
                            ** (1 / 5)
                        )
                        - 1
                    ) * 100
                else:
                    value = np.nan

        fcf_records.append(
            {
                "company_id": company_id,
                "fcf_cagr_5yr": value
            }
        )

    fcf_cagr_df = pd.DataFrame(
        fcf_records
    )

    # --------------------------------------------------------
    # MERGE EXTRA METRICS
    # --------------------------------------------------------

    df = base_df.merge(
        latest_roce,
        on="company_id",
        how="left"
    )

    df = df.merge(
        cfo_pat_df,
        on="company_id",
        how="left"
    )

    df = df.merge(
        fcf_cagr_df,
        on="company_id",
        how="left"
    )

    df["fcf_positive_flag"] = np.where(
        df["free_cash_flow_cr"] > 0,
        100.0,
        0.0
    )

    return df


# ============================================================
# COMPONENT SCORES
# ============================================================

def calculate_component_scores(df):
    df = df.copy()

    # Profitability
    df["roe_score"] = winsor_scale(
        df["return_on_equity_pct"]
    )

    df["roce_score"] = winsor_scale(
        df["computed_roce_pct"]
    )

    df["npm_score"] = winsor_scale(
        df["net_profit_margin_pct"]
    )

    # Cash quality
    df["fcf_cagr_score"] = winsor_scale(
        df["fcf_cagr_5yr"]
    )

    df["cfo_pat_score"] = winsor_scale(
        df["cfo_pat_ratio_5yr"]
    )

    # Growth
    df["revenue_growth_score"] = winsor_scale(
        df["revenue_cagr_5yr"]
    )

    df["pat_growth_score"] = winsor_scale(
        df["pat_cagr_5yr"]
    )

    # Leverage
    df["de_score"] = winsor_scale(
        df["debt_to_equity"],
        higher_is_better=False
    )

    df["icr_score"] = winsor_scale(
        df["effective_icr"]
        .replace(
            [np.inf, -np.inf],
            np.nan
        )
    )

    # Debt-free companies receive maximum ICR score
    debt_free = (
        df["total_debt_cr"].fillna(0)
        == 0
    )

    df.loc[
        debt_free,
        "icr_score"
    ] = 100.0

    return df


# ============================================================
# COMPOSITE SCORE
# ============================================================

def calculate_composite_score(df):
    df = calculate_component_scores(df)

    # Missing component scores use neutral 50
    score_columns = [
        "roe_score",
        "roce_score",
        "npm_score",
        "fcf_cagr_score",
        "cfo_pat_score",
        "revenue_growth_score",
        "pat_growth_score",
        "de_score",
        "icr_score"
    ]

    for col in score_columns:
        df[col] = df[col].fillna(50.0)

    df["global_composite_score"] = (
        df["roe_score"] * 0.15
        +
        df["roce_score"] * 0.10
        +
        df["npm_score"] * 0.10
        +
        df["fcf_cagr_score"] * 0.15
        +
        df["cfo_pat_score"] * 0.10
        +
        df["fcf_positive_flag"] * 0.05
        +
        df["revenue_growth_score"] * 0.10
        +
        df["pat_growth_score"] * 0.10
        +
        df["de_score"] * 0.10
        +
        df["icr_score"] * 0.05
    )

    df["global_composite_score"] = (
        df["global_composite_score"]
        .clip(0, 100)
        .round(2)
    )

    return df


# ============================================================
# SECTOR-RELATIVE SCORE
# ============================================================

def calculate_sector_relative_score(df):
    df = df.copy()

    def sector_scale(series):
        if len(series) == 1:
            return pd.Series(
                [50.0],
                index=series.index
            )

        return winsor_scale(series)

    df["sector_relative_score"] = (
        df
        .groupby(
            "broad_sector",
            group_keys=False
        )["global_composite_score"]
        .apply(sector_scale)
        .round(2)
    )

    # Use the new score as the official composite score
    df["composite_quality_score"] = (
        df["sector_relative_score"]
    )

    return df


# ============================================================
# PRESET THRESHOLDS FOR EXCEL COLOUR CODING
# ============================================================

PRESET_RULES = {
    "Quality Compounder": {
        "return_on_equity_pct": (">", 15),
        "debt_to_equity": ("<", 1.0),
        "free_cash_flow_cr": (">", 0),
        "revenue_cagr_5yr": (">", 10),
    },

    "Value Pick": {
        "pe_ratio": ("<", 20),
        "pb_ratio": ("<", 3.0),
        "debt_to_equity": ("<", 2.0),
        "dividend_yield_pct": (">", 1),
    },

    "Growth Accelerator": {
        "pat_cagr_5yr": (">", 20),
        "revenue_cagr_5yr": (">", 15),
        "debt_to_equity": ("<", 2.0),
    },

    "Dividend Champion": {
        "dividend_yield_pct": (">", 2),
        "dividend_payout_ratio_pct": ("<", 80),
        "free_cash_flow_cr": (">", 0),
    },

    "Debt-Free Blue Chip": {
        "debt_to_equity": ("==", 0),
        "return_on_equity_pct": (">", 12),
        "sales": (">", 5000),
    },

    "Turnaround Watch": {
        "revenue_cagr_3yr": (">", 10),
        "free_cash_flow_cr": (">", 0),
        "debt_to_equity_declining": ("==", True),
    },
}


# ============================================================
# 20 KPI EXPORT COLUMNS
# ============================================================

EXPORT_COLUMNS = [
    "company_id",
    "year",
    "broad_sector",
    "return_on_equity_pct",
    "computed_roce_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "fcf_cagr_5yr",
    "cfo_pat_ratio_5yr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "market_cap_crore",
    "global_composite_score",
    "sector_relative_score",
    "composite_quality_score",
]


# ============================================================
# APPLY EXCEL COLOUR FORMATTING
# ============================================================

def apply_excel_formatting(path, preset_results):
    wb = load_workbook(path)

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    for preset_name, result in preset_results.items():
        sheet_name = preset_name[:31]

        ws = wb[sheet_name]

        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center"
            )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_map = {
            cell.value: cell.column
            for cell in ws[1]
        }

        rules = PRESET_RULES.get(
            preset_name,
            {}
        )

        for column_name, rule in rules.items():
            if column_name not in header_map:
                continue

            col_number = header_map[column_name]

            operator, threshold = rule

            for row_number in range(
                2,
                ws.max_row + 1
            ):
                cell = ws.cell(
                    row=row_number,
                    column=col_number
                )

                value = cell.value

                if value is None:
                    cell.fill = red_fill
                    continue

                if operator == ">":
                    passed = value > threshold

                elif operator == "<":
                    passed = value < threshold

                elif operator == "==":
                    passed = value == threshold

                else:
                    passed = False

                cell.fill = (
                    green_fill
                    if passed
                    else red_fill
                )

        # Column widths
        for column_cells in ws.columns:
            max_length = 0

            column_letter = (
                column_cells[0].column_letter
            )

            for cell in column_cells:
                value = cell.value

                if value is not None:
                    max_length = max(
                        max_length,
                        len(str(value))
                    )

            ws.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                28
            )

    wb.save(path)


# ============================================================
# RUN PRESETS USING SCORED DATA
# ============================================================

def generate_preset_results(df):
    results = {}

    results["Quality Compounder"] = (
        quality_compounder(df)
    )

    results["Value Pick"] = (
        value_pick(df)
    )

    results["Growth Accelerator"] = (
        growth_accelerator(df)
    )

    results["Dividend Champion"] = (
        dividend_champion(df)
    )

    results["Debt-Free Blue Chip"] = (
        debt_free_blue_chip(df)
    )

    results["Turnaround Watch"] = (
        turnaround_watch(
            df,
            DB_PATH
        )
    )

    return results


# ============================================================
# EXPORT WORKBOOK
# ============================================================

def export_workbook(results):
    os.makedirs(
        "output",
        exist_ok=True
    )

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl"
    ) as writer:

        for preset_name, result in results.items():
            result = result.sort_values(
                "composite_quality_score",
                ascending=False,
                na_position="last"
            )

            columns = [
                col
                for col in EXPORT_COLUMNS
                if col in result.columns
            ]

            # Include turnaround-specific fields
            if (
                preset_name == "Turnaround Watch"
                and
                "revenue_cagr_3yr"
                in result.columns
            ):
                columns.append(
                    "revenue_cagr_3yr"
                )

            if (
                preset_name == "Turnaround Watch"
                and
                "debt_to_equity_declining"
                in result.columns
            ):
                columns.append(
                    "debt_to_equity_declining"
                )

            result[
                columns
            ].to_excel(
                writer,
                sheet_name=preset_name[:31],
                index=False
            )

    apply_excel_formatting(
        OUTPUT_PATH,
        results
    )


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70)
    print("COMPOSITE SCORE & SCREENER EXPORT")
    print("=" * 70)

    print("\nLoading 92-company universe...")

    base_df = load_screener_dataframe(
        DB_PATH
    )

    print(
        f"Companies loaded: {len(base_df)}"
    )

    print("\nBuilding additional scoring metrics...")

    df = build_extra_metrics(
        base_df
    )

    print("Calculating P10/P90 component scores...")

    df = calculate_composite_score(
        df
    )

    print("Calculating sector-relative scores...")

    df = calculate_sector_relative_score(
        df
    )

    print("Running six presets...")

    results = generate_preset_results(
        df
    )

    for name, result in results.items():
        print(
            f"{name:<25}: {len(result)} companies"
        )

    print("\nGenerating Excel workbook...")

    export_workbook(
        results
    )

    print(
        f"\nCreated: {OUTPUT_PATH}"
    )

    print("\nScore range:")

    print(
        "Global:",
        round(
            df["global_composite_score"].min(),
            2
        ),
        "to",
        round(
            df["global_composite_score"].max(),
            2
        )
    )

    print(
        "Sector Relative:",
        round(
            df["sector_relative_score"].min(),
            2
        ),
        "to",
        round(
            df["sector_relative_score"].max(),
            2
        )
    )

    print("\nProcessing complete.")


if __name__ == "__main__":
    main()