import os
import sqlite3

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from src.screener.engine import load_screener_dataframe
from src.screener.scoring_export import (
    build_extra_metrics,
    calculate_composite_score,
    calculate_sector_relative_score,
)


DB_PATH = "database/stock_analysis.db"
PEER_FILE = "data/peer_groups.xlsx"
OUTPUT_FILE = "output/peer_comparison.xlsx"


# ============================================================
# 20 REPORT METRICS
# ============================================================

METRICS = [
    "return_on_equity_pct",
    "computed_roce_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "fcf_cagr_5yr",
    "cfo_pat_ratio_5yr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "earnings_per_share",
    "book_value_per_share",
    "dividend_payout_ratio_pct",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "composite_quality_score",
]


# Lower value is better
INVERSE_METRICS = {
    "debt_to_equity",
    "pe_ratio",
    "pb_ratio",
}


# ============================================================
# LOAD PEER GROUPS
# ============================================================

def load_peer_groups():
    peers = pd.read_excel(
        PEER_FILE,
        sheet_name="Sheet1"
    )

    required = {
        "company_id",
        "peer_group_name",
        "is_benchmark"
    }

    missing = required - set(peers.columns)

    if missing:
        raise ValueError(
            f"Missing peer-group columns: {missing}"
        )

    peers["company_id"] = (
        peers["company_id"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    return peers


# ============================================================
# FIND COMPANY NAME
# ============================================================

def load_company_names():
    """
    Try database tables first.
    If no usable company_name column exists,
    company_id is used as fallback.
    """

    conn = sqlite3.connect(DB_PATH)

    names = None

    try:
        tables = pd.read_sql_query(
            """
            SELECT name
            FROM sqlite_master
            WHERE type = 'table'
            """,
            conn
        )["name"].tolist()

        for table in tables:
            try:
                columns = pd.read_sql_query(
                    f"PRAGMA table_info('{table}')",
                    conn
                )["name"].tolist()

                if (
                    "company_id" in columns
                    and
                    "company_name" in columns
                ):
                    names = pd.read_sql_query(
                        f"""
                        SELECT DISTINCT
                            company_id,
                            company_name
                        FROM "{table}"
                        """,
                        conn
                    )

                    break

            except Exception:
                continue

    finally:
        conn.close()

    if names is None:
        return pd.DataFrame(
            columns=[
                "company_id",
                "company_name"
            ]
        )

    names["company_id"] = (
        names["company_id"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    names = names.drop_duplicates(
        "company_id"
    )

    return names


# ============================================================
# LOAD 92-COMPANY METRIC DATA
# ============================================================

def load_metric_data():
    print("Loading company metrics...")

    df = load_screener_dataframe(
        DB_PATH
    )

    df = build_extra_metrics(df)

    df = calculate_composite_score(df)

    df = calculate_sector_relative_score(df)

    df["company_id"] = (
        df["company_id"]
        .astype(str)
        .str.strip()
        .str.upper()
    )

    names = load_company_names()

    if not names.empty:
        df = df.merge(
            names,
            on="company_id",
            how="left"
        )

    if "company_name" not in df.columns:
        df["company_name"] = df["company_id"]

    df["company_name"] = (
        df["company_name"]
        .fillna(df["company_id"])
    )

    return df


# ============================================================
# PERCENT RANK
# ============================================================

def percent_rank(series, inverse=False):
    """
    SQL-style PERCENT_RANK:
    (rank - 1) / (n - 1)

    Returns percentile values from 0 to 1.
    """

    numeric = pd.to_numeric(
        series,
        errors="coerce"
    )

    valid_count = numeric.notna().sum()

    if valid_count == 0:
        return pd.Series(
            np.nan,
            index=series.index
        )

    if valid_count == 1:
        result = pd.Series(
            np.nan,
            index=series.index
        )

        result.loc[
            numeric.notna()
        ] = 1.0

        return result

    ranks = numeric.rank(
        method="min",
        ascending=True
    )

    result = (
        (ranks - 1)
        /
        (valid_count - 1)
    )

    if inverse:
        result = 1 - result

    return result


# ============================================================
# BUILD ONE PEER GROUP SHEET
# ============================================================

def build_peer_group_sheet(
    peer_group_name,
    peer_mapping,
    metric_data
):
    group_mapping = peer_mapping[
        peer_mapping["peer_group_name"]
        == peer_group_name
    ].copy()

    group = group_mapping.merge(
        metric_data,
        on="company_id",
        how="left"
    )

    # --------------------------------------------------------
    # Ensure all 20 metrics exist
    # --------------------------------------------------------

    for metric in METRICS:
        if metric not in group.columns:
            group[metric] = np.nan

    # --------------------------------------------------------
    # Calculate percentile for every metric
    # --------------------------------------------------------

    percentile_columns = []

    for metric in METRICS:
        percentile_column = (
            f"{metric}_percentile"
        )

        group[percentile_column] = (
            percent_rank(
                group[metric],
                inverse=(
                    metric in INVERSE_METRICS
                )
            )
        )

        percentile_columns.append(
            percentile_column
        )

    # --------------------------------------------------------
    # Select output columns
    # --------------------------------------------------------

    output_columns = [
        "company_id",
        "company_name",
    ]

    for metric in METRICS:
        output_columns.append(metric)
        output_columns.append(
            f"{metric}_percentile"
        )

    output_columns.append(
        "is_benchmark"
    )

    result = group[
        output_columns
    ].copy()

    return result


# ============================================================
# ADD MEDIAN SUMMARY ROW
# ============================================================

def add_median_row(df):
    median_row = {}

    for column in df.columns:

        if column == "company_id":
            median_row[column] = "PEER MEDIAN"

        elif column == "company_name":
            median_row[column] = "Peer Group Median"

        elif column == "is_benchmark":
            median_row[column] = False

        elif pd.api.types.is_numeric_dtype(
            df[column]
        ):
            median_row[column] = (
                df[column].median()
            )

        else:
            median_row[column] = None

    return pd.concat(
        [
            df,
            pd.DataFrame([median_row])
        ],
        ignore_index=True
    )


# ============================================================
# SAFE EXCEL SHEET NAME
# ============================================================

def safe_sheet_name(name):
    invalid = [
        "\\",
        "/",
        "*",
        "?",
        ":",
        "[",
        "]"
    ]

    result = str(name)

    for character in invalid:
        result = result.replace(
            character,
            "-"
        )

    return result[:31]


# ============================================================
# EXPORT RAW WORKBOOK
# ============================================================

def export_workbook():
    os.makedirs(
        "output",
        exist_ok=True
    )

    peers = load_peer_groups()

    metric_data = load_metric_data()

    peer_groups = (
        peers["peer_group_name"]
        .dropna()
        .unique()
        .tolist()
    )

    print(
        f"Peer groups found: {len(peer_groups)}"
    )

    sheet_data = {}

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        for peer_group in peer_groups:

            result = build_peer_group_sheet(
                peer_group,
                peers,
                metric_data
            )

            result = add_median_row(
                result
            )

            sheet_name = safe_sheet_name(
                peer_group
            )

            result.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

            sheet_data[sheet_name] = result

            print(
                f"{peer_group:<30} "
                f"{len(result) - 1} companies"
            )

    format_workbook(
        sheet_data
    )

    return sheet_data


# ============================================================
# FORMAT EXCEL WORKBOOK
# ============================================================

def format_workbook(sheet_data):
    workbook = load_workbook(
        OUTPUT_FILE
    )

    # Percentile colours
    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE"
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C"
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE"
    )

    # Benchmark row
    gold_fill = PatternFill(
        fill_type="solid",
        fgColor="FFD966"
    )

    # Median row
    median_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7"
    )

    # Header
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9E1F2"
    )

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        # ----------------------------------------------------
        # Header formatting
        # ----------------------------------------------------

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(
                bold=True
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = thin_border

        ws.freeze_panes = "C2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 45

        header_map = {
            cell.value: cell.column
            for cell in ws[1]
        }

        benchmark_column = header_map.get(
            "is_benchmark"
        )

        median_row_number = ws.max_row

        # ----------------------------------------------------
        # Percentile colour coding
        # ----------------------------------------------------

        for header, column_number in (
            header_map.items()
        ):

            if not str(header).endswith(
                "_percentile"
            ):
                continue

            for row_number in range(
                2,
                median_row_number
            ):
                cell = ws.cell(
                    row=row_number,
                    column=column_number
                )

                value = cell.value

                if value is None:
                    continue

                if value >= 0.75:
                    cell.fill = green_fill

                elif value <= 0.25:
                    cell.fill = red_fill

                else:
                    cell.fill = yellow_fill

                cell.number_format = "0.0%"

        # ----------------------------------------------------
        # Benchmark row highlighting
        # ----------------------------------------------------

        if benchmark_column:

            for row_number in range(
                2,
                median_row_number
            ):

                benchmark_value = ws.cell(
                    row=row_number,
                    column=benchmark_column
                ).value

                if benchmark_value in (
                    True,
                    1,
                    "TRUE",
                    "True"
                ):

                    for column_number in range(
                        1,
                        ws.max_column + 1
                    ):
                        ws.cell(
                            row=row_number,
                            column=column_number
                        ).fill = gold_fill

        # ----------------------------------------------------
        # Median row formatting
        # ----------------------------------------------------

        for cell in ws[
            median_row_number
        ]:
            cell.fill = median_fill
            cell.font = Font(
                bold=True
            )
            cell.border = thin_border

        # ----------------------------------------------------
        # General cell formatting
        # ----------------------------------------------------

        for row in ws.iter_rows(
            min_row=2
        ):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    vertical="center"
                )

        # ----------------------------------------------------
        # Column widths
        # ----------------------------------------------------

        for column_cells in ws.columns:

            column_letter = (
                column_cells[0]
                .column_letter
            )

            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            ws.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                24
            )

        # Hide technical benchmark column
        if benchmark_column:
            benchmark_letter = ws.cell(
                row=1,
                column=benchmark_column
            ).column_letter

            ws.column_dimensions[
                benchmark_letter
            ].hidden = True

    workbook.save(
        OUTPUT_FILE
    )


# ============================================================
# VALIDATION
# ============================================================

def validate_workbook():
    workbook = load_workbook(
        OUTPUT_FILE,
        read_only=True
    )

    print("\nValidation")
    print("=" * 70)

    print(
        "Sheets:",
        len(workbook.sheetnames)
    )

    print(
        "Sheet names:",
        workbook.sheetnames
    )

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]

        print(
            f"{sheet_name:<30} "
            f"Rows: {ws.max_row:<4} "
            f"Columns: {ws.max_column}"
        )

    workbook.close()


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70)
    print("PEER COMPARISON EXCEL REPORT")
    print("=" * 70)

    export_workbook()

    print(
        f"\nCreated: {OUTPUT_FILE}"
    )

    validate_workbook()

    print(
        "\nProcessing complete."
    )


if __name__ == "__main__":
    main()